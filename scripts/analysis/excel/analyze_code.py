import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns

# class for relevant area and index of worksheet thats being analyzed
class CheckerWorksheet:
    def __init__(self, start_row, end_row, start_col, end_col, j):
        self.start_row = start_row
        self.end_row = end_row
        self.start_col = start_col
        self.end_col = end_col
        self.j = j

# class for subtask with its amount of errors and the sheet index
class Subtask:
    def __init__(self, sum=0, formula=0, number=0, format=0, sheet_index=0):
        self.sum = sum
        self.formula = formula
        self.number = number
        self.format = format
        self.sheet_index = sheet_index

    # method for counting errors in a subtask
    def classify_error(self, start_row, end_row, start_col, end_col, wb, wb2, wb_sol, wb_sol2, j): 
        ws = wb.worksheets[j]
        ws_sol = wb_sol.worksheets[j]
        ws2 = wb2.worksheets[j]
        ws_sol2 = wb_sol2.worksheets[j]
        for row in range(start_row, end_row):
            for col in range(start_col, end_col):
                char = get_column_letter(col)
                if ws[char + str(row)].value != ws_sol[char + str(row)].value:
                    if ws_sol[char + str(row)].value != ws_sol2[char + str(row)].value:
                        self.sum += 1
                        self.formula += 1
                    if ws_sol[char + str(row)].value == ws_sol2[char + str(row)].value and ws2[char + str(row)].value != ws_sol2[char + str(row)].value:
                        self.sum += 1
                        self.number += 1
                if ws[char + str(row)].font != ws_sol[char + str(row)].font:
                    self.sum += 1
                    self.format += 1
        self.sheet_index = j
        return self.sum, self.formula, self.number, self.format, self.sheet_index
        

def analyze(path, path_sol):
    res = []
    path = Path(path)
    path_sol = Path(path_sol)
    file_count = 0
    total_sum = total_formula = total_number = total_format = 0
    overall = overall_formula = overall_number = overall_format = 0
   
    # counts ".xlsx"-files in directory, thats supposed to be analyzed
    for file in os.listdir(path):
        if (file.endswith("xlsx")):
            file_count += 1
    print('File count:', file_count)

    wb_sol = load_workbook(path_sol)
    wb_sol2 = load_workbook(path_sol, data_only=True)

    # creating worksheet objects with relevant values and inserting them into a list
    sheet_def = []
    sheet = CheckerWorksheet(1, 34, 2, 16, 1)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 37, 2, 16, 2)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 57, 2, 5, 3)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 89, 1, 3, 4)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 202, 1, 6, 5)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 37, 1, 6, 6)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 48, 1, 8, 7)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 10, 1, 5, 8)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 2323, 11, 13, 9)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 27, 1, 8, 10)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 37, 2, 16, 11) # Placeholder, sheet will be skipped: only graph to compare
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 41, 5, 12, 12)
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 10, 1, 5, 13) # Placeholder, sheet will be skipped: only graph to compare
    sheet_def.append(sheet)
    sheet = CheckerWorksheet(1, 10, 1, 5, 14) # Placeholder, sheet will be skipped: only graph to compare
    sheet_def.append(sheet)

    # looping over submissions in given directory and analysing them
    # results are stored in "subtask"-objects and inserted into a list
    for i, input in enumerate(path.glob("*.xlsx*")): # rglob if directory includes subfolders and you want to access those aswell
        wb = load_workbook(filename=input)
        wb2 = load_workbook(filename=input, data_only=True)
       
        print(i)
        if len(wb.worksheets) <= len(wb_sol.worksheets):
            for sheet in sheet_def:
                if (sheet.j < len(wb.worksheets) and sheet.j < len(wb_sol.worksheets)):
                    if sheet.j != 11 and sheet.j != 13 and sheet.j != 14:
                        subtask = Subtask()
                        subtask.classify_error(sheet.start_row, sheet.end_row, sheet.start_col, sheet.end_col, wb, wb2, wb_sol, wb_sol2, sheet.j)
                        res.append(subtask)
        
        else:
            print("Studentenabgabe hat zu viele Sheets!")


    # iterating over list of "subtask"-objects and counting errors of subtasks with the same index 
    n = 1
    while n < len(wb_sol.worksheets):
        for entry in res:
            if n == entry.sheet_index:
                total_sum += entry.sum
                total_formula += entry.formula
                total_number += entry.number
                total_format += entry.format

        print(f"In Unteraufgabe {n} wurden Insgesamt {total_sum} Fehler gemacht!")
        print(f"Davon waren {total_formula} Formel-, {total_number} Nummern- & {total_format} Formatfehler!")

        # visualisation of errortypes per subtask as pie chart
        if n != 11 and n != 13 and n != 14:
            data = [total_formula, total_number, total_format]
            labels = ["formula errors", "number errors" , "format errors"]

            colors = sns.color_palette('pastel')[0:5]
            plt.pie(data, labels = labels, colors = colors, autopct='%.0f%%')
            plt.show()

        n += 1
        overall += total_sum
        overall_formula += total_formula
        overall_number += total_number
        overall_format += total_format

        total_sum = 0
        total_formula = 0
        total_number = 0
        total_format = 0


    print(f"In allen Abgaben wurden somit insgesamt {overall} Fehler gemacht!") 
    print(f"Davon waren {overall_formula} Formel-, {overall_number} Nummern- & {overall_format} Formatfehler!")
    data = [overall_formula, overall_number, overall_format]
    labels = ["formula errors", "number errors" , "format errors"]

    # visualisation of errortypes for total errorcount of all submissions as pie chart
    colors = sns.color_palette('pastel')[0:5]
    plt.pie(data, labels = labels, colors = colors, autopct='%.0f%%')
    plt.show()