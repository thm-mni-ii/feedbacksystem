from pathlib import Path
from openpyxl import load_workbook

def anonymize(path):
    path = Path(path)
    for i, input in enumerate(path.glob("*.xlsx*")): # rglob if directory includes subfolders and you want to access those aswell
        wb = load_workbook(filename=input)
        ws = wb["1 - Stammblatt"]

        # Cells which include personal data and the values they're replaced with
        ws["B5"] = f"Benutzerkennung Anonym {i}"
        ws["B6"] = f"Nachname Anonym {i}"
        ws["B7"] = f"Vorname Anonym {i}"
        ws["B8"] = f"Geburtsdatum Anonym {i}"
        ws["B10"].hyperlink = None
        ws["B10"] = f"E-Mail Anonym {i}"

        # Saving the changes in workbook 
        wb.save(input)